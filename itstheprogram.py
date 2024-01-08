#!/bin/python3

import bs4
import json
import math
import re
import requests
import time
import traceback
import typing
import urllib.parse

import tqdm

import contextlib

# import itertools
# import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Quick persistence


@contextlib.contextmanager
def FiledJson(file_path, default={}):
    try:
        with open(file_path, 'r', encoding="utf-8") as fp:
            obj = json.load(fp)
    except FileNotFoundError:
        obj = default

    try:
        yield obj
    finally:
        with open(file_path, 'w', encoding="utf-8") as fp:
            json.dump(obj, fp)

# Initialize a selenium browser, but also load/save cookies and refresh access token.
def initBrowserAndCookies():
    browser = webdriver.Firefox()

    steam_history_url = 'https://store.steampowered.com/account/history/'
    browser.get(steam_history_url)
    with FiledJson("cookiestore.json") as cookiestore:
        for cookie in cookiestore.get('sel', []):
            browser.add_cookie(cookie)

    browser.get(steam_history_url)

    if browser.current_url != steam_history_url:
        print(f"Please log into Steam using the new firefox window.")
        print(f"Note: Logging into Steam like this grants the program full access over your account and is a major security risk! Consult the readme file for details.")
        print("If you get an 'infinte loop', delete the cookiestore.json file and try again.")

        browser.delete_all_cookies()
        # TODO: Detect this case and clear cookies

        print("Waiting for login...")
        WebDriverWait(browser, timeout=math.inf).until(lambda browser: browser.current_url == steam_history_url)

    print("Navigated to", steam_history_url)

    with FiledJson("cookiestore.json") as cookiestore:
        cookiestore['sel'] = browser.get_cookies()
        cookiestore['req'] = {c['name']: c['value'] for c in cookiestore['sel']}

        time.sleep(1)
        cookiestore['access_token'] = browser.execute_script("return (new URLSearchParams(new URL(window.performance.getEntries().map(e => e.name).filter(n => n.includes('access_token'))[0]).search)).get('access_token')")

    return browser


def getPurchaseHistory() -> list[dict]:
    browser = initBrowserAndCookies()

    with FiledJson("cookiestore.json") as cookiestore:
        cookies = cookiestore['req']

    # Load whole page
    while True:
        try:
            WebDriverWait(browser, timeout=4).until(EC.element_to_be_clickable((By.ID, "load_more_button"))).click()
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        except Exception:
            traceback.print_exc()
            break

    print("Loaded all pages")

    # TODO: Keep track of processed transactions and only process new entries

    rows = browser.find_elements(By.CLASS_NAME, 'wallet_table_row')
    transactions = []
    for row in rows:
        try:
            if row.get_attribute('onclick') == "location.href='https://steamcommunity.com/market/#myhistory'":
                # Skip market transactions
                continue
            transid = re.search(r'transid=([0-9]+)', row.get_attribute('onclick')).group(1)
            transactions.append(transid)
        except Exception:
            print(row)
            print(row.get_attribute('onclick'))
            traceback.print_exc()

    purchased_items: list[dict] = []

    # For each transaction id, get granular purchase data. (Doesn't need selenium)
    for transaction_id in tqdm.tqdm(transactions):
        try:
            purchased_items += list(purchaseDetailsFromWizard(transaction_id, cookies=cookies))
        except Exception:
            print(transaction_id)
            traceback.print_exc()
            continue

    return purchased_items


def purchaseDetailsFromWizard(transaction_id, cookies, browser=None) -> typing.Iterator[dict]:
    # Scrape purchase details by using the "help with transaction" wizard
    try:
        wizard_url = f"https://help.steampowered.com/en/wizard/HelpWithTransaction?transid={transaction_id}"
        if browser:
            browser.get(wizard_url)
        resp_trans = requests.get(wizard_url, cookies=cookies)
        resp_trans.raise_for_status()
        soup_trans = bs4.BeautifulSoup(resp_trans.text, features="lxml")

        # purchase_date = soup_trans.find(class_='purchase_date').text

        for line_item_button in soup_trans.select("a[href*='/en/wizard/HelpWithMyPurchase']"):

            line_item_url = line_item_button['href']

            if browser:
                browser.get(line_item_url)

            is_gift = False
            if len(line_item_button.select("img[src*='icon_gift.png']")) > 0:
                is_gift = True

            resp_product = requests.get(line_item_url, cookies=cookies)
            resp_product.raise_for_status()
            soup_product = bs4.BeautifulSoup(resp_product.text, features="lxml")

            product_appids = [
                urllib.parse.parse_qs(urllib.parse.urlparse(a['href']).query)['appid'][0]
                for a in
                soup_product.select("a[href*='/en/wizard/HelpWithGame/']")
            ]
            product_infotags = [
                ','.join(btn.text for btn in a.findAll(class_="help_wizard_button_dark"))
                for a in
                soup_product.select("a[href*='/en/wizard/HelpWithGame/']")
            ]
            # primary_name = soup_product.select("a[href*='/HelpWithGame/']")[0].text.strip()
            primary_name = soup_product.find(class_='purchase_detail_field').text

            entry_row = {
                "primary_name": primary_name,
                "value": soup_product.find(class_='refund_value').text,
                "purchase_date": soup_product.find(class_='purchase_date').text.replace('Purchased: ', ''),
                "transaction_id": transaction_id,
                "appids": ','.join(product_appids),
                "infotags": ','.join(product_infotags),
                "is_gift": is_gift
            }
            yield entry_row

    except Exception:
        print(wizard_url, line_item_url)
        raise


def getPlaytime(appids, playtime_data):
    appids = appids.split(' ')
    mapped = {
        str(game['appid']): game['playtime_forever']
        for game in playtime_data['response']['games']
    }
    return sum([mapped.get(appid, 0) for appid in appids])


def writePurchaseXls(purchase_history, playtime_data):
    from openpyxl import Workbook

    xls_filename = "purchases.xlsx"

    print(f"Writing data to spreadsheet {xls_filename!r}")

    wb = Workbook()

    ws = wb.active
    ws.title = "Steam"

    ws.append((
        "Name",
        "Price",
        "Purchase Date",
        "Transaction ID",
        "Gift?",
        "App IDs",
        "Total Playtime (minutes)",
        "Info tags",
        "Hours per Dollar"
    ))

    for row in purchase_history:
        ws.append((
            str(row['primary_name']),  # A
            float(row['value'].replace('$', '')),  # B
            row['purchase_date'],  # C
            int(row['transaction_id']),  # D
            row['is_gift'],  # E
            row['appids'],  # F
            getPlaytime(row['appids'], playtime_data),  # G
            row['infotags'],  # H
            ""  # I
        ))
    for cell in ws["B"]:
        cell.number_format = "$0.00"
    for j, cell in enumerate(ws["I"]):
        i = j + 1
        if i == 1:
            continue
        ws[f"I{i}"] = f"=(G{i}/60)/B{i}"

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['C'].hidden = True
    ws.column_dimensions['D'].hidden = True
    ws.column_dimensions['F'].hidden = True
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['I'].width = 18

    ws["J1"] = "Overall hours per dollar"
    ws["J2"] = "=(SUM(G:G)/60)/SUM(B:B)"

    # Save the file
    try:
        wb.save(xls_filename)
    except PermissionError:
        print(f"Couldn't save the output file {xls_filename!r}! This probably means you have it open in excel already. Excel is very protective of files it has open!")
        raise


if __name__ == "__main__":
    # Load cached purchase history if it exists.
    # Otherwise, scrape your transaction history.
    purchase_history_path = 'purchase_history.json'
    try:
        with open(purchase_history_path, 'r') as fp:
            purchase_history = json.load(fp)
            print(f"Loaded saved purchase history from {purchase_history_path!r}")

    except Exception:
        purchase_history = getPurchaseHistory()

        with open(purchase_history_path, 'w') as fp:
            json.dump(purchase_history, fp, indent=2)
            print(f"Saved scraped data to {purchase_history_path!r}")

    # Get the latest playtime infomation from the API.
    # If it fails the first time, try to refresh the API token using selenium.
    failures = 0
    while True:
        if failures > 1:
            initBrowserAndCookies()

        try:
            with FiledJson("cookiestore.json") as cookiestore:
                print(f"Getting the latest playtime data from the Steam API")
                endpoint_url = f"https://api.steampowered.com/IPlayerService/GetOwnedGames/v0001/?access_token={cookiestore['access_token']}&steamid={cookiestore['req']['steamLoginSecure'].split('%7C')[0]}&format=json"
                resp = requests.get(endpoint_url, cookies=cookiestore['req'])
                resp.raise_for_status()
                playtime_data = resp.json()
            break
        except Exception:
            traceback.print_exc()
            failures += 1
            if failures > 2:
                raise

    writePurchaseXls(purchase_history, playtime_data)

    print("Done!")
